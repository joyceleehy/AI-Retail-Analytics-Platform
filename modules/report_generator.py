import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
import io

# Reusable styling
HEADER_FILL = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="2E86AB")

def style_header_row(ws, row_num):
    """Apply bold white text on blue background to a header row"""
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

def auto_fit_columns(ws):
    """Adjust column widths based on content length, safely skipping merged cells"""
    column_widths = {}

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if cell.value is not None:
                col_idx = cell.column
                length = len(str(cell.value))
                column_widths[col_idx] = max(column_widths.get(col_idx, 0), length)

    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(width + 4, 40)

def add_dataframe_sheet(wb, sheet_name, title, df):
    """
    Creates a new sheet, adds a title row, then writes the DataFrame
    below it with styled headers.
    """
    ws = wb.create_sheet(sheet_name)

    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    style_header_row(ws, 3)
    auto_fit_columns(ws)

    return ws

def generate_excel_report(kpis, rfm_summary, forecast_df, rca_category_breakdown=None):
    """
    Main function: builds a multi-sheet Excel workbook combining
    KPIs, RFM segmentation, Sales Forecast, and optionally Root Cause data.

    Returns: BytesIO object (in-memory file, ready for Streamlit download)
    """
    wb = Workbook()
    wb.remove(wb.active)

    # --- Sheet 1: Executive Summary (KPIs) ---
    ws1 = wb.create_sheet("Executive Summary")
    ws1["A1"] = "Executive KPI Summary"
    ws1["A1"].font = TITLE_FONT

    kpi_rows = [
        ("Total Revenue", f"${kpis['total_revenue']:,.0f}"),
        ("Total Profit", f"${kpis['total_profit']:,.0f}"),
        ("YoY Growth Rate", f"{kpis['growth_rate']:.1f}%"),
        ("Active Customers", f"{kpis['active_customers']:,}"),
        ("Average Order Value", f"${kpis['avg_order_value']:,.2f}"),
    ]

    ws1["A3"] = "Metric"
    ws1["B3"] = "Value"
    style_header_row(ws1, 3)

    for i, (label, value) in enumerate(kpi_rows, start=4):
        ws1[f"A{i}"] = label
        ws1[f"B{i}"] = value

    auto_fit_columns(ws1)

    # --- Sheet 2: Customer Segmentation (RFM) ---
    add_dataframe_sheet(wb, "Customer Segmentation", "RFM Segment Summary", rfm_summary)

    # --- Sheet 3: Sales Forecast ---
    forecast_display = forecast_df.copy()
    forecast_display["Month"] = forecast_display["Month"].dt.strftime("%B %Y")
    add_dataframe_sheet(wb, "Sales Forecast", "3-Month Sales Forecast", forecast_display)

    # --- Sheet 4: Root Cause Analysis (optional) ---
    if rca_category_breakdown is not None:
        add_dataframe_sheet(wb, "Root Cause Analysis", "Category Performance Breakdown", rca_category_breakdown)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer